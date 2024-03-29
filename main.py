from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib


background = "#06283D"
framebg = "#ededed"
framefg = "#05283d"

root = Tk()
root.title("Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)
root.resizable(False, False)

file = pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "DOB"
    sheet['F1'] = "Date of Registration"
    sheet['G1'] = "Religion"
    sheet['H1'] = "Skill"
    sheet['I1'] = "Father's Name"
    sheet['J1'] = "Mother's Name"
    sheet['K1'] = "Father's Occupation"
    sheet['L1'] = "Mother's Occupation"

    file.save('Student_data.xlsx')


# Exit Window command
def exit_form():
    root.destroy()


# Upload Image command
def show_image():
    global filename
    global img2
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select Image File", filetypes=(("JPG File", "*.jpg"),
                                                                                                        ("PNG File", "*.png"),
                                                                                                        ("All Files", "*.txt")))
    img2 = (Image.open(filename))
    resized_image = img2.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2


# ######################Registration NO.####################################
# it is created to automatically enter registration no.

def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row, column=1).value

    try:
        Registration.set(max_row_value + 1)
    except:
        Registration.set(1)


# ######################Clear All Command###################33
def clear_all():
    global img
    Name.set("")
    DOB.set("")
    Religion.set("")
    Skill.set("")
    F_Name.set("")
    M_Name.set("")
    Father_Occupation.set("")
    Mother_Occupation.set("")
    Class.set("Select Class")

    registration_no()

    save_button.config(state="normal")

    img1 = PhotoImage(file='images/passport.png')
    lbl.config(image=img1)
    lbl.image = img1

    img = ""


# ########################Save Command############################
def save_data():
    global img2
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    try:
        G1 = gender
    except:
        messagebox.showerror("Error", "Select Gender!")
    D2 = DOB.get()
    D1 = Date.get()
    Rel = Religion.get()
    S1 = Skill.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()

    if N1 == "" or C1 == "Select Class" or D2 == "" or D1 == "" or Rel == "" or S1 == "" or fathername == "" or mothername == "" or \
            F1 == "" or M1 == "":
        messagebox.showerror("Error", "All fields must be filled!")
    else:
        file = openpyxl.load_workbook("Student_data.xlsx")
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row+1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=C1)
        sheet.cell(column=4, row=sheet.max_row, value=G1)
        sheet.cell(column=5, row=sheet.max_row, value=D2)
        sheet.cell(column=6, row=sheet.max_row, value=D1)
        sheet.cell(column=7, row=sheet.max_row, value=Rel)
        sheet.cell(column=8, row=sheet.max_row, value=S1)
        sheet.cell(column=9, row=sheet.max_row, value=fathername)
        sheet.cell(column=10, row=sheet.max_row, value=mothername)
        sheet.cell(column=11, row=sheet.max_row, value=F1)
        sheet.cell(column=12, row=sheet.max_row, value=M1)
        file.save(r"Student_data.xlsx")

        try:
            img2.save("Student Images/" + str(R1) + ".jpg")
        except:
            messagebox.showinfo("Info", "Profile Picture is not available!!!!")

        messagebox.showinfo("Info", "Successfully data entered!!!")

        # clear entry boxes and image
        clear_all()
        # update registration number
        registration_no()


# #######################Search Command############################
def search():
    global img3
    text = Search.get()

    clear_all()
    save_button.config(state='disabled')

    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]
            reg_number = str(name)[15: -1]
    try:
        print(name)
    except:
        messagebox.showerror("Invalid Reg", "Invalid Registration Number!!!")

    x1 = sheet.cell(row=int(reg_number), column=1).value
    x2 = sheet.cell(row=int(reg_number), column=2).value
    x3 = sheet.cell(row=int(reg_number), column=3).value
    x4 = sheet.cell(row=int(reg_number), column=4).value
    x5 = sheet.cell(row=int(reg_number), column=5).value
    x6 = sheet.cell(row=int(reg_number), column=6).value
    x7 = sheet.cell(row=int(reg_number), column=7).value
    x8 = sheet.cell(row=int(reg_number), column=8).value
    x9 = sheet.cell(row=int(reg_number), column=9).value
    x10 = sheet.cell(row=int(reg_number), column=10).value
    x11 = sheet.cell(row=int(reg_number), column=11).value
    x12 = sheet.cell(row=int(reg_number), column=12).value

    Registration.set(x1)
    Name.set(x2)
    Class.set(x3)
    if x4 == "Female":
        R2.select()
    else:
        R1.select()
    DOB.set(x5)
    Date.set(x6)
    Religion.set(x7)
    Skill.set(x8)
    F_Name.set(x9)
    M_Name.set(x10)
    Father_Occupation.set(x11)
    Mother_Occupation.set(x12)
    try:
        img3 = (Image.open("Student Images/" + str(x1) + ".jpg"))
        resized_image = img3.resize((190, 190))
        photo2 = ImageTk.PhotoImage(resized_image)
        lbl.config(image=photo2)
        lbl.image = photo2
    except:
        pass


# ##########################Update Command#############################
def update():
    global img3
    global img2
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    selection()
    G1 = gender
    D2 = DOB.get()
    D1 = Date.get()
    Rel = Religion.get()
    S1 = Skill.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()

    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == R1:
            name = row[0]
            reg_number = str(name)[15: -1]

    if N1 == "" or C1 == "Select Class" or D2 == "" or D1 == "" or Rel == "" or S1 == "" or fathername == "" or mothername == "" or \
            F1 == "" or M1 == "":
        messagebox.showerror("Error", "All fields must be filled!")
    else:
        sheet.cell(column=1, row=int(reg_number), value=R1)
        sheet.cell(column=2, row=int(reg_number), value=N1)
        sheet.cell(column=3, row=int(reg_number), value=C1)
        sheet.cell(column=4, row=int(reg_number), value=G1)
        sheet.cell(column=5, row=int(reg_number), value=D2)
        sheet.cell(column=6, row=int(reg_number), value=D1)
        sheet.cell(column=7, row=int(reg_number), value=Rel)
        sheet.cell(column=8, row=int(reg_number), value=S1)
        sheet.cell(column=9, row=int(reg_number), value=fathername)
        sheet.cell(column=10, row=int(reg_number), value=mothername)
        sheet.cell(column=11, row=int(reg_number), value=F1)
        sheet.cell(column=12, row=int(reg_number), value=M1)
        file.save(r'Student_data.xlsx')

        try:
            img3.save("Student Images/" + str(R1) + ".jpg")
        except:
            img2.save("Student Images/" + str(R1) + ".jpg")

        messagebox.showinfo("Update", "Information Updated Successfully!!")
        # clear all entry boxes after update
        clear_all()


# gender selection
def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
    else:
        gender = "Female"


# top frames
Label(root, text="Email: nemesai@gmail.com", width=10, height=3, bg="#cdd4d8", anchor='e').pack(side=TOP, fill=X)
Label(root, text="STUDENT REGISTRATION", width=10, height=2, bg="#506877", fg="#333300", font='georgia 20 bold').pack(
    side=TOP, fill=X)

# search box and button
Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=3, font="georgia 20").place(x=820, y=65)
imageicon3 = PhotoImage(file="images/search.png")
Srch = Button(root, text="Search", compound=LEFT, image=imageicon3, width=123, bg="#68ddfa", font="georgia 13 bold", command=search)
Srch.place(x=1060, y=66)

# update button
imageicon4 = PhotoImage(file="images/switch.png")
update_button = Button(root, image=imageicon4, bg="#cdd4d8", command=update)
update_button.place(x=110, y=64)

# registration and date
Label(root, text="Registration No:", font="georgia 13", fg=framebg, bg=background).place(x=30, y=150)
Label(root, text="Date:", font="georgia 13", fg=framebg, bg=background).place(x=500, y=150)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font="georgia 12", state="readonly")
reg_entry.place(x=170, y=150)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font="georgia 12", state="readonly")
date_entry.place(x=550, y=150)

Date.set(d1)

# Student details
obj = LabelFrame(root, text="Student's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text="Full Name:", font="Georgia 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj, text="Date of Birth:", font="georgia 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj, text="Gender:", font="georgia 13", bg=framebg, fg=framefg).place(x=30, y=150)

Label(obj, text="Class:", font="georgia 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj, text="Religion:", font="georgia 13", bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj, text="Skills:", font="georgia 13", bg=framebg, fg=framefg).place(x=500, y=150)

Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=30, font="georgia 12")
name_entry.place(x=160, y=50)

DOB = StringVar()
dob_entry = Entry(obj, textvariable=DOB, width=17, font="georgia 12")
dob_entry.place(x=160, y=100)

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=150, y=150)

R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=250, y=150)

Religion = StringVar()
religion_entry = Entry(obj, textvariable=Religion, width=15, font="georgia 12")
religion_entry.place(x=630, y=100)

Skill = StringVar()
skill_entry = Entry(obj, textvariable=Skill, width=15, font="georgia 12")
skill_entry.place(x=630, y=150)

Class = Combobox(obj, values=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12'], font="georgia 10", width=15, state="r")
Class.place(x=630, y=50)
Class.set("Select Class")

# Parents details
obj2 = LabelFrame(root, text="Parent's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=220, relief=GROOVE)
obj2.place(x=30, y=470)

Label(obj2, text="Father's Name:", font="georgia 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj2, text="Occupation:", font="georgia 13", bg=framebg, fg=framefg).place(x=30, y=100)

F_Name = StringVar()
f_entry = Entry(obj2, textvariable=F_Name, width=30, font="georgia 12")
f_entry.place(x=160, y=50)

Father_Occupation = StringVar()
FO_entry = Entry(obj2, textvariable=Father_Occupation, width=20, font="georgia 12")
FO_entry.place(x=160, y=100)

Label(obj2, text="Mother's Name:", font="georgia 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj2, text="Occupation:", font="georgia 13", bg=framebg, fg=framefg).place(x=500, y=100)

M_Name = StringVar()
m_entry = Entry(obj2, textvariable=M_Name, width=25, font="georgia 12")
m_entry.place(x=640, y=50)

Mother_Occupation = StringVar()
MO_entry = Entry(obj2, textvariable=Mother_Occupation, width=20, font="georgia 12")
MO_entry.place(x=630, y=100)

# image box
f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)

img = PhotoImage(file="images/passport.png")
lbl = Label(f, bg="black", image=img)
lbl.place(x=0, y=0)

# buttons
Button(root, text="Upload", width=19, height=2, font="georgia 12 bold", bg="lightblue", command=show_image).place(x=1000, y=370)
save_button = Button(root, text="Save", width=19, height=2, font="georgia 12 bold", bg="lightgreen", command=save_data)
save_button.place(x=1000, y=450)
Button(root, text="Reset", width=19, height=2, font="georgia 12 bold", bg="lightpink", command=clear_all).place(x=1000, y=530)
Button(root, text="Exit", width=19, height=2, font="georgia 12 bold", bg="grey", command=exit_form).place(x=1000, y=610)


root.mainloop()
